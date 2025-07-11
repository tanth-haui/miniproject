from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

SIGNAL_MAP = {
    "Actual Speed": "vNE",
    "Set Speed": "bvNSET0",
    "Feed Forward": "vQLDAC",
    "AC Switch": "vSWMONT"
}

def get_csv_files(input_path: Path) -> list:
    if not input_path.is_dir():
        raise FileNotFoundError(f"❌ Thư mục input không tồn tại: {input_path}")
    csv_files = list(input_path.glob("*.csv"))
    if not csv_files:
        raise FileNotFoundError("⚠ Không tìm thấy file .csv nào trong thư mục input.")
    return csv_files

def ensure_output_folder(base_output_path: Path) -> Path:
    if not base_output_path.is_dir():
        raise FileNotFoundError(f"❌ Đường dẫn output không tồn tại: {base_output_path}")
    output_path = base_output_path if base_output_path.name.lower() == "output" else base_output_path / "output"
    if not output_path.exists():
        try:
            output_path.mkdir(parents=True)
        except Exception as error:
            raise PermissionError(f"❌ Không thể tạo thư mục output: {error}")
        
    #kiểm tra folder output có thể tạo file được không
    try:
        test_file = output_path / "test_write.tmp"
        with open(test_file, "w") as f:
            f.write("test")
        test_file.unlink()
    except Exception as error:
        raise PermissionError(f"❌ Không thể ghi vào thư mục output: {error}")
    return output_path

def parse_time(start_str, end_str):
    try:
        if start_str.strip().startswith("-") or end_str.strip().startswith("-"):
            raise ValueError("❌ Start Time và End Time không được là số âm.")
        start_time = int(start_str)
        end_time = int(end_str)
        if start_time < 0 or end_time < 0:
            raise ValueError("❌ Thời gian phải là số nguyên không âm.")
        if end_time <= start_time:
            raise ValueError("❌ Start Time phải nhỏ hơn End Time.")
        return start_time, end_time
    except ValueError as value_error:
        if "invalid literal" in str(value_error):
            raise ValueError("❌ Giá trị thời gian phải là số nguyên hợp lệ.")
        raise

def find_column(df, signal_key):
    for col in df.columns:
        base = col.split("\\")[0].strip()
        if base == signal_key:
            return col
    raise ValueError(f"⚠ Không tìm thấy cột có prefix \"{signal_key}\".")

def create_plot(df, time_col, value_col, filename_stem: str, output_dir: Path) -> Path:
    plt.figure(figsize=(8, 5))
    plt.plot(df[time_col], df[value_col], color="green", linewidth=1.5, label=value_col)
    plt.title(filename_stem, fontsize=11)
    # plt.grid(True, which="both", linestyle="--", linewidth=0.5)
    plt.tight_layout()
    plt.legend(fontsize=9, loc="best")
    plot_path = output_dir / f"{filename_stem}.png"
    plt.savefig(plot_path, dpi=800)
    plt.close()
    return plot_path

def create_output_csv(file_path, df_filtered, output_folder: Path):
    output_file = output_folder / f"{file_path.stem}.csv"
    df_filtered.to_csv(output_file, index=False, encoding="utf-8-sig")
    print(f"✅ Đã tạo file: {output_file.name}")

def cleanup_png_files(folder: Path):
    png_files = list(folder.glob("*.png"))
    for file in png_files:
        try:
            file.unlink()
            print(f"🧹 Đã xóa ảnh: {file.name}")
        except Exception as error:
            print(f"⚠ Không thể xóa {file.name}: {error}")

def create_summary_excel(results: dict, start_time, end_time, signal, output_folder: Path):
    summary_path = output_folder / "summary.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    font_bold = Font(name="Arial", size=11, bold=True)
    font_regular = Font(name="Arial", size=11)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")

    for idx, (file_path, result) in enumerate(results.items()):
        col_offset = idx * 10
        base_col = 1 + col_offset
        labels = ["File Name(s):", "Time Range:", "Signal:"]
        values = [file_path.stem, f"{start_time} - {end_time}", signal]

        for i, label in enumerate(labels, start=1):
            label_cell = ws.cell(row=i, column=base_col, value=label)
            label_cell.font = font_bold
            label_cell.alignment = align_center
            ws.merge_cells(start_row=i, start_column=base_col + 1, end_row=i, end_column=base_col + 9)
            value_cell = ws.cell(row=i, column=base_col + 1, value=values[i - 1])
            value_cell.font = font_regular
            value_cell.alignment = align_left

        for col in range(base_col, base_col + 10):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 14 if col == base_col else 10

        for row_i in range(1, 4):
            for col_j in range(base_col, base_col + 10):
                ws.cell(row=row_i, column=col_j).border = Border(top=thin, bottom=thin, left=thin, right=thin)

        for row in range(1, 37):
            ws.cell(row=row, column=base_col + 10).border = Border(left=thick)
        for col in range(base_col, base_col + 10):
            ws.cell(row=37, column=col).border = Border(top=thick)

        chart_path = result["plot"]
        if chart_path.exists():
            img = XLImage(str(chart_path))
            img.width, img.height = 700, 680
            img_cell = f"{get_column_letter(base_col + 1)}6"
            ws.add_image(img, img_cell)

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 18.75

    wb.save(summary_path)
    print(f"📄 Đã tạo file summary: {summary_path.name}")
    print(f"🖼️ Đã đưa vào Excel {len(results)} ảnh.")

def run_processing(input_folder, base_output_folder, signal_selection, start_time_str, end_time_str):
    input_path = Path(input_folder)
    base_output_path = Path(base_output_folder)

    # Kiểm tra input folder và file CSV
    csv_files = get_csv_files(input_path)

    # Kiểm tra tín hiệu
    signal_prefix = SIGNAL_MAP.get(signal_selection)
    if not signal_prefix:
        raise ValueError("❌ Tín hiệu không hợp lệ.")

    # Kiểm tra thời gian
    start, end = parse_time(start_time_str, end_time_str)

    # Chỉ tạo thư mục output sau khi tất cả đầu vào hợp lệ
    output_path = ensure_output_folder(base_output_path)

    results = {}
    for file_path in csv_files:
        try:
            df = pd.read_csv(file_path)
            time_col = df.columns[0]
            signal_col = find_column(df, signal_prefix)
            df_filtered = df[[time_col, signal_col]]
            df_filtered = df_filtered[(df_filtered[time_col] >= start) & (df_filtered[time_col] <= end)]

            if df_filtered.empty:
                print(f"⚠ {file_path.name}: Không có dòng nào thỏa mãn.")
                continue

            create_output_csv(file_path, df_filtered, output_path)
            plot_path = create_plot(df_filtered, time_col, signal_col, file_path.stem, output_path)
            print(f"📊 Đã tạo biểu đồ: {plot_path.name}")

            results[file_path] = {
                "df": df_filtered,
                "plot": plot_path
            }
        except Exception as error:
            print(f"❌ Lỗi khi xử lý {file_path.name}: {error}")

    if results:
        create_summary_excel(results, start, end, signal_prefix, output_path)
        cleanup_png_files(output_path)
