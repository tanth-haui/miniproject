from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# Cấu hình
INPUT_FOLDER = Path(r"D:\Code_learn\Mini_test\info\input")
OUTPUT_FOLDER = Path(r"D:\Code_learn\Mini_test\output")
SIGNAL_SELECTION = "Feed Forward"
START_TIME = "0"
END_TIME = "200000"

SIGNAL_MAP = {
    "Actual Speed": "vNE",
    "Set Speed": "bvNSET0",
    "Feed Forward": "vQLDAC",
    "AC Switch": "vSWMONT"
}

def validate_and_prepare():
    if not INPUT_FOLDER.is_dir():
        raise FileNotFoundError(f"❌ Thư mục input không tồn tại: {INPUT_FOLDER}")
    csv_files = list(INPUT_FOLDER.glob("*.csv"))
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    return csv_files

def parse_time(start_str, end_str):
    try:
        if start_str.strip().startswith("-") or end_str.strip().startswith("-"):
            raise ValueError("❌ Start Time và End Time không được là số âm.")
        start_time = int(start_str)
        end_time = int(end_str)
        if start_time < 0 or end_time < 0:
            raise ValueError("❌ Thời gian phải là số nguyên không âm.")
        if end_time <= start_time:
            raise ValueError("❌ Start Time phải nhỏ hơn End Time.")
        return start_time, end_time
    except ValueError as ve:
        if "invalid literal" in str(ve):
            raise ValueError("❌ Giá trị thời gian phải là số nguyên hợp lệ.")
        raise

def find_column(df, signal_key):
    for col in df.columns:
        base = col.split("\\")[0].strip()
        if base == signal_key:
            return col
    raise ValueError(f"⚠ Không tìm thấy cột có prefix '{signal_key}'.")

def create_plot(df, time_col, value_col, filename_stem: str, output_dir: Path) -> Path:
    plt.figure(figsize=(6, 4))
    plt.plot(df[time_col], df[value_col], color="green", linewidth=1.5)
    plt.title(filename_stem, fontsize=11)
    plt.grid(True, which='both', linestyle='--', linewidth=0.5)
    plt.tight_layout()
    plot_path = output_dir / f"{filename_stem}.png"
    plt.savefig(plot_path)
    plt.close()
    return plot_path

def create_summary_excel(results: dict):
    summary_path = OUTPUT_FOLDER / "summary.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    font_bold = Font(name="Arial", size=11, bold=True)
    font_regular = Font(name="Arial", size=11)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")

    for idx, (file_path, result) in enumerate(results.items()):
        col_offset = idx * 10
        base_col = 1 + col_offset

        labels = ["File Name(s):", "Time Range:", "Signal:"]
        values = [file_path.stem, f"{START_TIME} → {END_TIME}", SIGNAL_SELECTION]

        for i, label in enumerate(labels, start=1):
            label_cell = ws.cell(row=i, column=base_col, value=label)
            label_cell.font = font_bold
            label_cell.alignment = align_center
            ws.merge_cells(start_row=i, start_column=base_col + 1, end_row=i, end_column=base_col + 9)
            value_cell = ws.cell(row=i, column=base_col + 1, value=values[i - 1])
            value_cell.font = font_regular
            value_cell.alignment = align_left

        for col in range(base_col, base_col + 10):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 14 if col == base_col else 10

        for r in range(1, 4):
            for c in range(base_col, base_col + 10):
                ws.cell(row=r, column=c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
        for row in range(1, 37):
            ws.cell(row=row, column=base_col + 10).border = Border(left=thick)
        for col in range(base_col, base_col + 10):
            ws.cell(row=37, column=col).border = Border(top=thick)

        chart_path = result["plot"]
        if chart_path.exists():
            img = XLImage(str(chart_path))
            img.width, img.height = 590, 570
            img_cell = f"{get_column_letter(base_col + 1)}6"
            ws.add_image(img, img_cell)

    wb.save(summary_path)
    print(f"📄 Đã tạo file summary: {summary_path.name}")
    print(f"🖼️ Đã nhúng {len(results)} ảnh vào file Excel.")

def cleanup_png_files(folder: Path):
    png_files = list(folder.glob("*.png"))
    for file in png_files:
        try:
            file.unlink()
            print(f"🧹 Đã xóa ảnh: {file.name}")
        except Exception as e:
            print(f"⚠ Không thể xóa {file.name}: {e}")

def create_output_csv(file_path, df_filtered):
    output_file = OUTPUT_FOLDER / f"{file_path.stem}.csv"
    df_filtered.to_csv(output_file, index=False, encoding="utf-8-sig")
    print(f"✅ Đã tạo file: {output_file.name}")

def main():
    try:
        files = validate_and_prepare()
        if not files:
            print("⚠ Không tìm thấy file .csv nào trong thư mục input.")
            return

        start, end = parse_time(START_TIME, END_TIME)
        signal_prefix = SIGNAL_MAP.get(SIGNAL_SELECTION)
        if not signal_prefix:
            raise ValueError("Tín hiệu không hợp lệ.")

        results = {}
        for file_path in files:
            try:
                df = pd.read_csv(file_path)
                time_col = df.columns[0]
                signal_col = find_column(df, signal_prefix)
                df_filtered = df[[time_col, signal_col]]
                df_filtered = df_filtered[
                    (df_filtered[time_col] >= start) & (df_filtered[time_col] <= end)
                ]
                if df_filtered.empty:
                    print(f"⚠ {file_path.name}: Không có dòng nào thỏa mãn.")
                    continue

                create_output_csv(file_path, df_filtered)
                plot_path = create_plot(df_filtered, time_col, signal_col, file_path.stem, OUTPUT_FOLDER)
                print(f"📊 Đã tạo biểu đồ: {plot_path.name}")

                results[file_path] = {
                    "df": df_filtered,
                    "plot": plot_path
                }

            except Exception as e:
                print(f"❌ Lỗi khi xử lý {file_path.name}: {e}")

        if results:
            create_summary_excel(results)
            cleanup_png_files(OUTPUT_FOLDER)

    except Exception as e:
        print(f"❌ Lỗi chính: {e}")
main()
